"""
Testes para o módulo de geração de Excel (excel.py).

Cobertura:
- create_excel() com diferentes inputs (vazio, único, múltiplos)
- Validação de formatação (cores, fontes, bordas)
- Parsing de datas em múltiplos formatos
- Metadados e fórmulas
- Edge cases (valores None, strings vazias)
"""

import logging
from contextlib import contextmanager
from datetime import datetime
from io import BytesIO

import pytest
from openpyxl import load_workbook

from excel import create_excel, parse_datetime, sanitize_for_excel, _validate_licitacoes_types


@contextmanager
def open_workbook(buffer: BytesIO):
    """Context manager para garantir fechamento do workbook."""
    wb = load_workbook(buffer)
    try:
        yield wb
    finally:
        wb.close()


class TestParseDateTime:
    """Testes para parse_datetime()."""

    def test_parse_iso_with_z_timezone(self):
        """Deve parsear ISO 8601 com Z (UTC)."""
        result = parse_datetime("2024-01-25T10:30:00Z")
        assert result == datetime(2024, 1, 25, 10, 30, 0)

    def test_parse_iso_with_offset_timezone(self):
        """Deve parsear ISO 8601 com offset (+00:00)."""
        result = parse_datetime("2024-01-25T10:30:00+00:00")
        assert result == datetime(2024, 1, 25, 10, 30, 0)

    def test_parse_iso_without_timezone(self):
        """Deve parsear ISO 8601 sem timezone."""
        result = parse_datetime("2024-01-25T10:30:00")
        assert result == datetime(2024, 1, 25, 10, 30, 0)

    def test_parse_date_only(self):
        """Deve parsear data simples (YYYY-MM-DD)."""
        result = parse_datetime("2024-01-25")
        assert result == datetime(2024, 1, 25, 0, 0, 0)

    def test_parse_none(self):
        """Deve retornar None para valor None."""
        assert parse_datetime(None) is None

    def test_parse_empty_string(self):
        """Deve retornar None para string vazia."""
        assert parse_datetime("") is None

    def test_parse_invalid_format(self):
        """Deve retornar None para formato inválido."""
        assert parse_datetime("25/01/2024") is None
        assert parse_datetime("not a date") is None


class TestCreateExcel:
    """Testes para create_excel()."""

    def test_create_excel_returns_bytesio(self):
        """Deve retornar BytesIO buffer."""
        result = create_excel([])
        assert isinstance(result, BytesIO)

    def test_create_excel_with_empty_list(self):
        """Deve gerar Excel válido mesmo com lista vazia."""
        buffer = create_excel([])

        with open_workbook(buffer) as wb:
            # Verificar sheet principal existe
            assert "Licitações Uniformes" in wb.sheetnames
            ws = wb["Licitações Uniformes"]

            # Verificar headers
            assert ws["A1"].value == "Código PNCP"
            assert ws["B1"].value == "Objeto"
            assert ws["K1"].value == "Link"

            # Verificar formatação do header
            assert ws["A1"].font.bold is True
            assert (
                ws["A1"].font.color.rgb == "00FFFFFF"
            )  # Branco (openpyxl ARGB format)
            assert (
                ws["A1"].fill.start_color.rgb == "002E7D32"
            )  # Verde (openpyxl ARGB format)

            # Verificar que não há linha de totais (lista vazia)
            assert ws["F3"].value is None

    def test_create_excel_with_single_item(self):
        """Deve gerar Excel com um item e linha de totais."""
        licitacao = {
            "codigoCompra": "12345",
            "numeroControlePNCP": "12345678000100-1-000123/2024",
            "objetoCompra": "Aquisição de uniformes escolares",
            "nomeOrgao": "Prefeitura Municipal",
            "uf": "SP",
            "municipio": "São Paulo",
            "valorTotalEstimado": 150000.50,
            "modalidadeNome": "Pregão Eletrônico",
            "dataPublicacaoPncp": "2024-01-20T08:00:00Z",
            "dataAberturaProposta": "2024-02-01T10:00:00Z",
            "situacaoCompraNome": "Em Disputa",
            "linkSistemaOrigem": "https://sistema.compras.gov.br/edital/12345",
        }

        buffer = create_excel([licitacao])

        with open_workbook(buffer) as wb:
            ws = wb["Licitações Uniformes"]

            # Verificar dados na linha 2
            assert ws["A2"].value == "12345"
            assert ws["B2"].value == "Aquisição de uniformes escolares"
            assert ws["C2"].value == "Prefeitura Municipal"
            assert ws["D2"].value == "SP"
            assert ws["E2"].value == "São Paulo"
            assert ws["F2"].value == 150000.50
            assert ws["G2"].value == "Pregão Eletrônico"
            assert ws["J2"].value == "Em Disputa"

            # Verificar formatação de moeda
            assert "R$" in ws["F2"].number_format

            # Verificar hyperlink (coluna K) - agora usando linkSistemaOrigem
            assert ws["K2"].value == "Abrir"
            assert ws["K2"].hyperlink.target == "https://sistema.compras.gov.br/edital/12345"
            assert ws["K2"].font.color.rgb == "000563C1"  # Azul (openpyxl ARGB format)
            assert ws["K2"].font.underline == "single"

            # Verificar linha de totais (row 3)
            assert ws["E3"].value == "TOTAL:"
            assert ws["E3"].font.bold is True
            assert "=SUM(F2:F2)" in ws["F3"].value
            assert ws["F3"].font.bold is True

            # Verificar bordas
            assert ws["A2"].border.left.style == "thin"

    def test_create_excel_with_multiple_items(self):
        """Deve gerar Excel com múltiplos itens."""
        licitacoes = [
            {
                "codigoCompra": "123",
                "objetoCompra": "Item 1",
                "valorTotalEstimado": 100000,
            },
            {
                "codigoCompra": "456",
                "objetoCompra": "Item 2",
                "valorTotalEstimado": 200000,
            },
            {
                "codigoCompra": "789",
                "objetoCompra": "Item 3",
                "valorTotalEstimado": 300000,
            },
        ]

        buffer = create_excel(licitacoes)

        with open_workbook(buffer) as wb:
            ws = wb["Licitações Uniformes"]

        # Verificar 3 linhas de dados
        assert ws["A2"].value == "123"
        assert ws["A3"].value == "456"
        assert ws["A4"].value == "789"

        # Verificar linha de totais (row 5)
        assert "=SUM(F2:F4)" in ws["F5"].value

    def test_create_excel_with_none_values(self):
        """Deve lidar com valores None nos campos."""
        licitacao = {
            "codigoCompra": "999",
            "objetoCompra": None,
            "valorTotalEstimado": None,
            "dataPublicacaoPncp": None,
            "dataAberturaProposta": None,
        }

        buffer = create_excel([licitacao])

        with open_workbook(buffer) as wb:
            ws = wb["Licitações Uniformes"]

        # Verificar que não crashou
        assert ws["A2"].value == "999"
        assert ws["B2"].value is None
        assert ws["F2"].value is None
        assert ws["H2"].value is None
        assert ws["I2"].value is None

    def test_create_excel_with_link_sistema_origem(self):
        """Deve usar linkSistemaOrigem quando disponível (prioridade 1)."""
        licitacao = {
            "numeroControlePNCP": "12345678000100-1-000001/2025",
            "linkSistemaOrigem": "https://sistema.compras.gov.br/edital/123",
            "linkProcessoEletronico": "https://processo.gov.br/456",
        }

        buffer = create_excel([licitacao])

        with open_workbook(buffer) as wb:
            ws = wb["Licitações Uniformes"]

        # Deve usar linkSistemaOrigem (prioridade)
        assert ws["K2"].hyperlink.target == "https://sistema.compras.gov.br/edital/123"

    def test_create_excel_with_link_processo_eletronico(self):
        """Deve usar linkProcessoEletronico quando linkSistemaOrigem não existe (prioridade 2)."""
        licitacao = {
            "numeroControlePNCP": "12345678000100-1-000001/2025",
            "linkProcessoEletronico": "https://processo.gov.br/456",
        }

        buffer = create_excel([licitacao])

        with open_workbook(buffer) as wb:
            ws = wb["Licitações Uniformes"]

        # Deve usar linkProcessoEletronico (segunda prioridade)
        assert ws["K2"].hyperlink.target == "https://processo.gov.br/456"

    def test_create_excel_with_fallback_link(self):
        """Deve gerar link padrão PNCP parseando numeroControlePNCP quando nenhum link específico existe."""
        licitacao = {"numeroControlePNCP": "12345678000100-1-000001/2025"}

        buffer = create_excel([licitacao])

        with open_workbook(buffer) as wb:
            ws = wb["Licitações Uniformes"]

        # Deve usar fallback parseando numeroControlePNCP: CNPJ/ANO/SEQUENCIAL
        # "12345678000100-1-000001/2025" -> /editais/12345678000100/2025/1
        assert (
            ws["K2"].hyperlink.target
            == "https://pncp.gov.br/app/editais/12345678000100/2025/1"
        )

    def test_create_excel_with_fallback_link_real_example(self):
        """Deve gerar link correto com exemplo real da API PNCP."""
        licitacao = {"numeroControlePNCP": "67366310000103-1-000189/2025"}

        buffer = create_excel([licitacao])

        with open_workbook(buffer) as wb:
            ws = wb["Licitações Uniformes"]

        # Exemplo real: "67366310000103-1-000189/2025" -> /editais/67366310000103/2025/189
        assert (
            ws["K2"].hyperlink.target
            == "https://pncp.gov.br/app/editais/67366310000103/2025/189"
        )

    def test_create_excel_with_invalid_numero_controle(self):
        """Deve usar busca genérica quando numeroControlePNCP tem formato inválido."""
        licitacao = {"numeroControlePNCP": "formato-invalido"}

        buffer = create_excel([licitacao])

        with open_workbook(buffer) as wb:
            ws = wb["Licitações Uniformes"]

        # Formato inválido -> busca genérica
        assert (
            ws["K2"].hyperlink.target
            == "https://pncp.gov.br/app/editais?q=formato-invalido"
        )

    def test_create_excel_frozen_panes(self):
        """Deve congelar a primeira linha (headers)."""
        buffer = create_excel([{"codigoCompra": "123"}])

        with open_workbook(buffer) as wb:
            ws = wb["Licitações Uniformes"]

            # Verificar frozen panes
            assert ws.freeze_panes == "A2"

    def test_create_excel_metadata_sheet(self):
        """Deve criar aba de Metadata com estatísticas."""
        licitacoes = [
            {"valorTotalEstimado": 100000},
            {"valorTotalEstimado": 200000},
            {"valorTotalEstimado": 300000},
        ]

        buffer = create_excel(licitacoes)

        with open_workbook(buffer) as wb:
            # Verificar que existe aba Metadata
            assert "Metadata" in wb.sheetnames
            ws_meta = wb["Metadata"]

            # Verificar conteúdo
            assert ws_meta["A1"].value == "Gerado em:"
            assert ws_meta["B1"].value is not None  # Timestamp

            assert ws_meta["A2"].value == "Total de licitações:"
            assert ws_meta["B2"].value == 3

            assert ws_meta["A3"].value == "Valor total estimado:"
            assert ws_meta["B3"].value == 600000
            assert "R$" in ws_meta["B3"].number_format

    def test_create_excel_metadata_with_none_values(self):
        """Metadata deve somar corretamente mesmo com None."""
        licitacoes = [
            {"valorTotalEstimado": 100000},
            {"valorTotalEstimado": None},  # Deve ignorar
            {"valorTotalEstimado": 50000},
        ]

        buffer = create_excel(licitacoes)

        with open_workbook(buffer) as wb:
            ws_meta = wb["Metadata"]

            assert ws_meta["B2"].value == 3  # 3 licitações
            assert ws_meta["B3"].value == 150000  # Soma ignora None

    def test_create_excel_date_formatting(self):
        """Deve aplicar formatação correta nas colunas de data."""
        licitacao = {
            "dataPublicacaoPncp": "2024-01-25T10:00:00Z",
            "dataAberturaProposta": "2024-02-01T14:30:00Z",
        }

        buffer = create_excel([licitacao])

        with open_workbook(buffer) as wb:
            ws = wb["Licitações Uniformes"]

            # Verificar formatação de data
            assert "DD/MM/YYYY" in ws["H2"].number_format  # Publicação
            assert "DD/MM/YYYY HH:MM" in ws["I2"].number_format  # Abertura

    def test_create_excel_column_widths(self):
        """Deve definir larguras corretas das colunas."""
        buffer = create_excel([])

        with open_workbook(buffer) as wb:
            ws = wb["Licitações Uniformes"]

            # Verificar algumas larguras (PRD Section 5.1)
            assert ws.column_dimensions["A"].width == 25  # Código PNCP
            assert ws.column_dimensions["B"].width == 60  # Objeto
            assert ws.column_dimensions["D"].width == 6  # UF

    def test_create_excel_invalid_input(self):
        """Deve lançar ValueError se input não for lista."""
        with pytest.raises(ValueError, match="licitacoes deve ser uma lista"):
            create_excel("not a list")

        with pytest.raises(ValueError):
            create_excel({"not": "a list"})

    def test_create_excel_can_be_reopened(self):
        """Buffer gerado deve ser um Excel válido que pode ser reaberto."""
        licitacao = {"codigoCompra": "TEST123", "objetoCompra": "Test"}
        buffer = create_excel([licitacao])

        # Salvar e reabrir para validar integridade
        with open_workbook(buffer) as wb:
            ws = wb["Licitações Uniformes"]

            assert ws["A2"].value == "TEST123"
            assert ws["B2"].value == "Test"

            # Verificar que pode ser salvo novamente
            new_buffer = BytesIO()
            wb.save(new_buffer)
            new_buffer.seek(0)

        with open_workbook(new_buffer) as wb2:
            assert "Licitações Uniformes" in wb2.sheetnames

    def test_create_excel_with_numero_controle_missing_parts(self):
        """Deve usar busca genérica quando numeroControlePNCP tem menos de 3 partes (sem tipo)."""
        # Edge case: linha 152 - len(cnpj_tipo_seq) < 3
        licitacao = {"numeroControlePNCP": "12345-000189/2025"}  # Missing tipo (-1-)

        buffer = create_excel([licitacao])

        with open_workbook(buffer) as wb:
            ws = wb["Licitações Uniformes"]

        # Deve cair no except e usar busca genérica
        assert (
            ws["K2"].hyperlink.target
            == "https://pncp.gov.br/app/editais?q=12345-000189/2025"
        )

    def test_create_excel_with_empty_cnpj_component(self):
        """Deve usar busca genérica quando CNPJ está vazio após parsing."""
        # Edge case: linha 160 - cnpj vazio
        licitacao = {"numeroControlePNCP": "-1-000189/2025"}  # Empty CNPJ

        buffer = create_excel([licitacao])

        with open_workbook(buffer) as wb:
            ws = wb["Licitações Uniformes"]

        # Deve cair no else (linha 160) e lançar ValueError -> busca genérica
        assert (
            ws["K2"].hyperlink.target
            == "https://pncp.gov.br/app/editais?q=-1-000189/2025"
        )

    def test_create_excel_with_empty_ano_component(self):
        """Deve usar busca genérica quando ano está vazio após parsing."""
        # Edge case: linha 160 - ano vazio
        licitacao = {"numeroControlePNCP": "12345678000100-1-000189/"}  # Empty ano

        buffer = create_excel([licitacao])

        with open_workbook(buffer) as wb:
            ws = wb["Licitações Uniformes"]

        # Deve cair no else (linha 160) e lançar ValueError -> busca genérica
        assert (
            ws["K2"].hyperlink.target
            == "https://pncp.gov.br/app/editais?q=12345678000100-1-000189/"
        )

    def test_create_excel_with_empty_sequencial_component(self):
        """Deve usar busca genérica quando sequencial é apenas zeros (lstrip retorna vazio)."""
        # Edge case: linha 160 - sequencial vazio após lstrip("0")
        licitacao = {"numeroControlePNCP": "12345678000100-1-000000/2025"}  # All zeros

        buffer = create_excel([licitacao])

        with open_workbook(buffer) as wb:
            ws = wb["Licitações Uniformes"]

        # sequencial.lstrip("0") = "" -> if cnpj and ano and sequencial fails -> linha 160
        assert (
            ws["K2"].hyperlink.target
            == "https://pncp.gov.br/app/editais?q=12345678000100-1-000000/2025"
        )

    def test_create_excel_with_large_dataset(self):
        """Deve gerar Excel com dataset grande (1000+ linhas) sem erros."""
        # Edge case: stress test com 1000 licitações
        licitacoes = [
            {
                "codigoCompra": f"CODE{i:04d}",
                "objetoCompra": f"Objeto da licitação {i}" * 10,  # Strings longas
                "nomeOrgao": f"Órgão {i}",
                "uf": "SP" if i % 2 == 0 else "RJ",
                "municipio": f"Município {i}",
                "valorTotalEstimado": 50000 + (i * 1000),
                "modalidadeNome": "Pregão Eletrônico",
                "dataPublicacaoPncp": f"2024-01-{(i % 28) + 1:02d}T08:00:00Z",
                "dataAberturaProposta": f"2024-02-{(i % 28) + 1:02d}T10:00:00Z",
                "situacaoCompraNome": "Em Disputa",
                "linkSistemaOrigem": f"https://sistema.compras.gov.br/edital/{i}",
            }
            for i in range(1000)
        ]

        buffer = create_excel(licitacoes)

        with open_workbook(buffer) as wb:
            ws = wb["Licitações Uniformes"]

            # Verificar primeira e última linha
            assert ws["A2"].value == "CODE0000"
            assert ws["A1001"].value == "CODE0999"  # Row 1001 = header + 1000 rows

            # Verificar linha de totais (row 1002)
            assert ws["E1002"].value == "TOTAL:"
            assert "=SUM(F2:F1001)" in ws["F1002"].value

            # Verificar metadata
            ws_meta = wb["Metadata"]
            assert ws_meta["B2"].value == 1000

    def test_create_excel_with_special_characters(self):
        """Deve lidar com caracteres especiais em todos os campos."""
        licitacao = {
            "codigoCompra": "ABC-123/2024",
            "objetoCompra": 'Aquisição de "uniformes" & <equipamentos> (diversos)',
            "nomeOrgao": "Prefeitura de São José dos Pinhais",
            "uf": "PR",
            "municipio": "São José dos Pinhais",
            "valorTotalEstimado": 123456.78,
            "modalidadeNome": "Pregão Eletrônico SRP",
            "situacaoCompraNome": "Pré-Publicação",
            "numeroControlePNCP": "12345678000100-1-000001/2024",
        }

        buffer = create_excel([licitacao])

        with open_workbook(buffer) as wb:
            ws = wb["Licitações Uniformes"]

            # Verificar que caracteres especiais foram preservados
            assert ws["A2"].value == "ABC-123/2024"
            assert '"uniformes"' in ws["B2"].value
            assert "&" in ws["B2"].value
            assert "São José" in ws["C2"].value
            assert ws["D2"].value == "PR"

    def test_create_excel_with_extremely_long_objeto(self):
        """Deve lidar com strings muito longas no campo objeto."""
        # Edge case: strings gigantes (>1000 chars)
        licitacao = {
            "codigoCompra": "LONG001",
            "objetoCompra": "A" * 5000,  # 5000 caracteres
            "valorTotalEstimado": 100000,
        }

        buffer = create_excel([licitacao])

        with open_workbook(buffer) as wb:
            ws = wb["Licitações Uniformes"]

            # Verificar que foi armazenado (Excel suporta até 32,767 chars)
            assert ws["B2"].value == "A" * 5000
            assert len(ws["B2"].value) == 5000

    def test_create_excel_with_zero_value(self):
        """Deve formatar corretamente valor zero."""
        licitacao = {
            "codigoCompra": "ZERO001",
            "valorTotalEstimado": 0.0,  # Valor zero
        }

        buffer = create_excel([licitacao])

        with open_workbook(buffer) as wb:
            ws = wb["Licitações Uniformes"]

            # Verificar formatação de moeda com valor zero
            assert ws["F2"].value == 0.0
            assert "R$" in ws["F2"].number_format

    def test_create_excel_with_negative_value(self):
        """Deve lidar com valores negativos (edge case improvável mas possível)."""
        licitacao = {
            "codigoCompra": "NEG001",
            "valorTotalEstimado": -50000.0,  # Valor negativo
        }

        buffer = create_excel([licitacao])

        with open_workbook(buffer) as wb:
            ws = wb["Licitações Uniformes"]

            # Verificar que aceita valor negativo
            assert ws["F2"].value == -50000.0
            assert "R$" in ws["F2"].number_format

    def test_create_excel_metadata_with_empty_list(self):
        """Metadata deve exibir zeros quando lista está vazia."""
        buffer = create_excel([])

        with open_workbook(buffer) as wb:
            ws_meta = wb["Metadata"]

            # Verificar valores zero
            assert ws_meta["B2"].value == 0  # Total de licitações
            assert ws_meta["B3"].value == 0  # Valor total
            assert "R$" in ws_meta["B3"].number_format

    def test_create_excel_with_illegal_control_characters(self):
        """
        Deve sanitizar caracteres de controle ilegais (bug fix).

        PNCP data sometimes contains \x13 (Device Control 3) instead of
        em-dash characters. This was causing openpyxl.IllegalCharacterError.

        Real error from production:
        "Ambulatório Médico de Especialidades \x13 AME SUL"
        """
        # Real text that caused the production error (with \x13 char)
        illegal_text = (
            "Aquisição de equipamentos para Ambulatório \x13 AME SUL, "
            "do Consórcio \x13 COMESP"
        )

        licitacao = {
            "codigoCompra": "TEST-ILLEGAL-001",
            "objetoCompra": illegal_text,
            "nomeOrgao": "Órgão com \x00 NUL e \x1f controle",
            "uf": "PR",
            "municipio": "Curitiba",
            "valorTotalEstimado": 500000.0,
            "situacaoCompraNome": "Em andamento \x0b com tab vertical",
        }

        # Should NOT raise IllegalCharacterError
        buffer = create_excel([licitacao])

        with open_workbook(buffer) as wb:
            ws = wb["Licitações Uniformes"]

            # Verify data was written (chars replaced with space)
            assert "Ambulatório" in ws["B2"].value
            assert "AME SUL" in ws["B2"].value
            # The \x13 should have been replaced
            assert "\x13" not in ws["B2"].value

            # Verify other sanitized fields
            assert "\x00" not in ws["C2"].value
            assert "\x1f" not in ws["C2"].value
            assert "\x0b" not in ws["J2"].value


class TestSanitizeForExcel:
    """Testes para sanitize_for_excel()."""

    def test_sanitize_removes_device_control_chars(self):
        """Deve remover caracteres Device Control (\x13, \x14, etc)."""
        text = "AME \x13 SUL \x14 NORTE"
        result = sanitize_for_excel(text)
        assert result == "AME   SUL   NORTE"
        assert "\x13" not in result
        assert "\x14" not in result

    def test_sanitize_removes_nul_character(self):
        """Deve remover caractere NUL (\x00)."""
        text = "Texto\x00com\x00nulos"
        result = sanitize_for_excel(text)
        assert result == "Texto com nulos"
        assert "\x00" not in result

    def test_sanitize_removes_all_illegal_chars(self):
        """Deve remover todos os caracteres ilegais do openpyxl."""
        # All illegal chars: \x00-\x08, \x0b-\x0c, \x0e-\x1f
        text = "A\x00B\x01C\x08D\x0bE\x0cF\x0eG\x1fH"
        result = sanitize_for_excel(text)
        assert result == "A B C D E F G H"

    def test_sanitize_preserves_allowed_whitespace(self):
        """Deve preservar tab (\x09), LF (\x0a), CR (\x0d)."""
        text = "Linha1\tcom\ttab\nLinha2\rLinha3"
        result = sanitize_for_excel(text)
        assert result == text  # Unchanged

    def test_sanitize_handles_none(self):
        """Deve retornar string vazia para None."""
        result = sanitize_for_excel(None)
        assert result == ""

    def test_sanitize_handles_empty_string(self):
        """Deve retornar string vazia para string vazia."""
        result = sanitize_for_excel("")
        assert result == ""

    def test_sanitize_handles_non_string(self):
        """Deve converter não-strings para string."""
        result = sanitize_for_excel(12345)
        assert result == "12345"

    def test_sanitize_preserves_unicode(self):
        """Deve preservar caracteres Unicode válidos."""
        text = "Licitação de móveis — R$ 50.000,00"
        result = sanitize_for_excel(text)
        assert result == text  # Unchanged

    def test_sanitize_real_production_error(self):
        """Teste com o texto exato que causou erro em produção."""
        # Text from production log (2026-02-05)
        text = (
            "Aquisição e instalação de equipamentos de informática, "
            "infraestrutura de rede lógica, rede sem fio de alta performance, "
            "equipamentos de videoconferência e projeção multimídia, bem como "
            "o respectivo mobiliário técnico (rack de piso), visando garantir "
            "a plena operação do novo prédio do Ambulatório Médico de "
            "Especialidades \x13 AME SUL, do Consórcio Metropolitano de "
            "Serviços do Paraná \x13 COMESP"
        )
        result = sanitize_for_excel(text)

        # Should not contain the illegal char
        assert "\x13" not in result

        # Should still contain the important text
        assert "AME SUL" in result
        assert "COMESP" in result
        assert "Ambulatório" in result


class TestValidateLicitacoesTypes:
    """Testes para _validate_licitacoes_types() — #180 TD-HP-003."""

    def test_no_warning_for_safe_types(self, caplog):
        """Não deve logar warning quando todos os valores são tipos seguros."""
        licitacoes = [
            {
                "codigoCompra": "123",
                "objetoCompra": "Uniformes",
                "valorTotalEstimado": 50000.0,
                "uf": None,
                "status": True,
            }
        ]
        with caplog.at_level(logging.WARNING, logger="excel"):
            _validate_licitacoes_types(licitacoes)

        assert "unexpected_type" not in caplog.text

    def test_warning_emitted_for_unexpected_type(self, caplog):
        """Deve logar warning quando um dict contém tipo não-serializável."""
        class CustomObj:
            pass

        licitacoes = [
            {
                "codigoCompra": "123",
                "metadados": CustomObj(),
            }
        ]
        with caplog.at_level(logging.WARNING, logger="excel"):
            _validate_licitacoes_types(licitacoes)

        assert "unexpected_type" in caplog.text

    def test_create_excel_with_unexpected_type_logs_warning_and_succeeds(self, caplog):
        """create_excel deve logar warning para tipo inesperado mas ainda retornar BytesIO válido.

        openpyxl/sanitize_for_excel coerce via str(), então geração continua.
        """
        class CustomObj:
            def __str__(self):
                return "custom-value"

        licitacao = {
            "codigoCompra": "TEST-001",
            "objetoCompra": "Materiais",
            "metadados": CustomObj(),
        }
        with caplog.at_level(logging.WARNING, logger="excel"):
            result = create_excel([licitacao])

        # Warning was logged
        assert "unexpected_type" in caplog.text
        # Generation still succeeded — result is a valid BytesIO
        assert isinstance(result, BytesIO)
        assert result.read(4) == b"PK\x03\x04"  # ZIP magic (xlsx)
